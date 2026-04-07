from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file
from flask_login import login_required, current_user
from models import Question, User, ExamAttempt
from extensions import db
from functools import wraps
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from datetime import datetime

admin_bp = Blueprint('admin', __name__)

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            flash('Bu sayfaya erişim yetkiniz yok.', 'danger')
            return redirect(url_for('main.dashboard'))
        return f(*args, **kwargs)
    return decorated_function

@admin_bp.route('/')
@login_required
@admin_required
def admin_panel():
    total_users = User.query.count()
    total_exams = ExamAttempt.query.filter_by(completed=True).count()
    total_questions = Question.query.count()
    
    recent_exams = ExamAttempt.query.filter_by(completed=True).order_by(
        ExamAttempt.end_time.desc()
    ).limit(10).all()
    
    return render_template('admin/panel.html',
                         total_users=total_users,
                         total_exams=total_exams,
                         total_questions=total_questions,
                         recent_exams=recent_exams)

@admin_bp.route('/questions')
@login_required
@admin_required
def questions():
    all_questions = Question.query.all()
    return render_template('admin/questions.html', questions=all_questions)

@admin_bp.route('/add-question', methods=['POST'])
@login_required
@admin_required
def add_question():
    question_text = request.form.get('question_text')
    question_type = request.form.get('question_type')
    difficulty = request.form.get('difficulty')
    points = int(request.form.get('points', 10))
    
    question = Question(
        question_text=question_text,
        question_type=question_type,
        difficulty=difficulty,
        points=points
    )
    
    if question_type == 'multiple_choice':
        question.option_a = request.form.get('option_a')
        question.option_b = request.form.get('option_b')
        question.option_c = request.form.get('option_c')
        question.option_d = request.form.get('option_d')
        question.correct_answer = request.form.get('correct_answer')
    
    db.session.add(question)
    db.session.commit()
    
    flash('Soru başarıyla eklendi!', 'success')
    return redirect(url_for('admin.questions'))

@admin_bp.route('/generate-questions', methods=['POST'])
@login_required
@admin_required
def generate_questions():
    from utils.ai_generator import generate_questions_with_ai
    
    topic = request.form.get('topic', 'Python programlama')
    difficulty = request.form.get('difficulty', 'intermediate')
    count = int(request.form.get('count', 5))
    
    try:
        questions = generate_questions_with_ai(topic, difficulty, count)
        flash(f'{len(questions)} soru başarıyla oluşturuldu!', 'success')
    except Exception as e:
        flash(f'Soru oluşturulurken hata: {str(e)}', 'danger')
    
    return redirect(url_for('admin.questions'))

@admin_bp.route('/exam-results')
@login_required
@admin_required
def exam_results():
    # Get filter parameters
    level_filter = request.args.get('level', 'all')
    user_filter = request.args.get('user', '')
    
    # Base query
    query = ExamAttempt.query.filter_by(completed=True)
    
    # Apply filters
    if level_filter != 'all':
        query = query.filter_by(exam_level=level_filter)
    
    if user_filter:
        query = query.join(User).filter(User.username.contains(user_filter))
    
    # Get results ordered by date
    exam_attempts = query.order_by(ExamAttempt.end_time.desc()).all()
    
    # Calculate statistics
    stats = {
        'total': ExamAttempt.query.filter_by(completed=True).count(),
        'none': ExamAttempt.query.filter_by(completed=True, exam_level='none').count(),
        'beginner': ExamAttempt.query.filter_by(completed=True, exam_level='beginner').count(),
        'intermediate': ExamAttempt.query.filter_by(completed=True, exam_level='intermediate').count(),
        'advanced': ExamAttempt.query.filter_by(completed=True, exam_level='advanced').count(),
        'avg_score': db.session.query(db.func.avg(ExamAttempt.score)).filter_by(completed=True).scalar() or 0
    }
    
    return render_template('admin/exam_results.html',
                         exam_attempts=exam_attempts,
                         stats=stats,
                         level_filter=level_filter,
                         user_filter=user_filter)

@admin_bp.route('/exam-detail/<int:attempt_id>')
@login_required
@admin_required
def exam_detail(attempt_id):
    exam_attempt = ExamAttempt.query.get_or_404(attempt_id)
    return render_template('admin/exam_detail.html', exam_attempt=exam_attempt)

@admin_bp.route('/users')
@login_required
@admin_required
def users():
    # Get filter parameters
    search = request.args.get('search', '')
    level_filter = request.args.get('level', 'all')
    admin_filter = request.args.get('admin', 'all')
    
    # Base query
    query = User.query
    
    # Apply filters
    if search:
        query = query.filter(
            (User.username.contains(search)) |
            (User.email.contains(search)) |
            (User.phone_number.contains(search))
        )
    
    if level_filter != 'all':
        query = query.filter_by(coding_knowledge=level_filter)
    
    if admin_filter == 'admin':
        query = query.filter_by(is_admin=True)
    elif admin_filter == 'user':
        query = query.filter_by(is_admin=False)
    
    # Get users ordered by creation date
    users_list = query.order_by(User.created_at.desc()).all()
    
    # Calculate statistics
    stats = {
        'total': User.query.count(),
        'admins': User.query.filter_by(is_admin=True).count(),
        'users': User.query.filter_by(is_admin=False).count(),
        'with_google': User.query.filter(User.google_id.isnot(None)).count(),
    }
    
    return render_template('admin/users.html',
                         users=users_list,
                         stats=stats,
                         search=search,
                         level_filter=level_filter,
                         admin_filter=admin_filter)

@admin_bp.route('/users/export')
@login_required
@admin_required
def export_users():
    # Get all users
    users = User.query.order_by(User.created_at.desc()).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Kullanıcılar"
    
    # Style definitions
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ['ID', 'Kullanıcı Adı', 'E-posta', 'Telefon', 'Kayıt Tarihi', 
               'Admin', 'Google OAuth', 'Kodlama Bilgisi', 'En İyi Puan', 'Seviye']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data rows
    for row_num, user in enumerate(users, 2):
        ws.cell(row=row_num, column=1, value=user.id)
        ws.cell(row=row_num, column=2, value=user.username)
        ws.cell(row=row_num, column=3, value=user.email)
        ws.cell(row=row_num, column=4, value=user.phone_number or '-')
        ws.cell(row=row_num, column=5, value=user.created_at.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row_num, column=6, value='Evet' if user.is_admin else 'Hayır')
        ws.cell(row=row_num, column=7, value='Evet' if user.google_id else 'Hayır')
        ws.cell(row=row_num, column=8, value=user.coding_knowledge)
        ws.cell(row=row_num, column=9, value=f"{user.get_best_score():.1f}")
        ws.cell(row=row_num, column=10, value=user.get_level())
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    filename = f"kullanicilar_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@admin_bp.route('/exam-results/export')
@login_required
@admin_required
def export_exam_results():
    # Get all completed exams
    exams = ExamAttempt.query.filter_by(completed=True).order_by(ExamAttempt.end_time.desc()).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sınav Sonuçları"
    
    # Style definitions
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ['ID', 'Kullanıcı Adı', 'E-posta', 'Telefon', 'Başlangıç', 'Bitiş', 
               'Süre (dk)', 'Puan (%)', 'Seviye', 'Sınav Seviyesi', 'Sekme Değiştirme']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data rows
    for row_num, exam in enumerate(exams, 2):
        duration = (exam.end_time - exam.start_time).seconds // 60 if exam.end_time else 0
        
        ws.cell(row=row_num, column=1, value=exam.id)
        ws.cell(row=row_num, column=2, value=exam.user.username)
        ws.cell(row=row_num, column=3, value=exam.user.email)
        ws.cell(row=row_num, column=4, value=exam.user.phone_number or '-')
        ws.cell(row=row_num, column=5, value=exam.start_time.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row_num, column=6, value=exam.end_time.strftime('%d/%m/%Y %H:%M') if exam.end_time else '-')
        ws.cell(row=row_num, column=7, value=duration)
        ws.cell(row=row_num, column=8, value=f"{exam.score:.1f}")
        ws.cell(row=row_num, column=9, value=exam.get_level())
        ws.cell(row=row_num, column=10, value=exam.exam_level)
        ws.cell(row=row_num, column=11, value=exam.tab_switches)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    filename = f"sinav_sonuclari_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@admin_bp.route('/users/toggle-admin/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def toggle_admin(user_id):
    user = User.query.get_or_404(user_id)
    
    if user.id == current_user.id:
        flash('Kendi yetkinizi değiştiremezsiniz!', 'danger')
    else:
        user.is_admin = not user.is_admin
        db.session.commit()
        flash(f'{user.username} kullanıcısı {"admin yapıldı" if user.is_admin else "kullanıcı yapıldı"}!', 'success')
    
    return redirect(url_for('admin.users'))

@admin_bp.route('/users/delete/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    
    if user.id == current_user.id:
        flash('Kendi hesabınızı silemezsiniz!', 'danger')
    else:
        username = user.username
        db.session.delete(user)
        db.session.commit()
        flash(f'{username} kullanıcısı silindi!', 'success')
    
    return redirect(url_for('admin.users'))
